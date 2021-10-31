import openpyxl  # 方便调用
from openpyxl import Workbook

# 创建工作簿对象
work_book = openpyxl.Workbook()

# 创建表<对象>
sheet = work_book.create_sheet('表1')

"""数据写入"""
for i in range(1, 10):
    for j in range(1, i + 1):
        print(f'{j} * {i} = {j * i}', end='\t')
        sheet.cell(row=i, column=j).value = f'{j} * {i} = {j * i}'
    print()

# 保存此工作蒲对象
work_book.save('实例.xlsx')

