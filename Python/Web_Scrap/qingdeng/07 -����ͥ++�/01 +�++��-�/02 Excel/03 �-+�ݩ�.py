import openpyxl  # 方便调用
from openpyxl import Workbook

# 根据excel文加名字读取
work_book = openpyxl.load_workbook('实例.xlsx')

# 获取所有表的名字
# print(work_book.sheetnames)

sheet = work_book['表1']

# 读取表1最大行数和列数
# print(sheet.max_row)  # 最大行
# print(sheet.max_column)  # 最大列

# 获取第一行数据
# for i in range(1, sheet.max_column + 1):
#     print(sheet.cell(row=1, column=i).value)  # 如果单元格为空则返回None

# # 获取第一列数据
# for i in range(1, sheet.max_row + 1):
#     print(sheet.cell(row=i, column=1).value)  # 如果单元格为空则返回None
for i in range(1, sheet.max_column + 1):
    for j in range(1, sheet.max_row + 1):
        print(sheet.cell(row=i, column=j).value)
