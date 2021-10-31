import openpyxl  # 方便调用
from openpyxl import Workbook

# 创建工作簿对象
work_book = openpyxl.Workbook()

# 创建表<对象>
sheet = work_book.create_sheet('表1')

"""数据写入"""
# sheet['A1'] = 'A1'
# sheet['B1'] = 'B1'

# cell()  是单元格对象, row表示单元格的行数, column 表示单元格的列数
# sheet.cell(row=1, column=1).value = '11111'
# sheet.cell(row=2, column=2).value = '222222'

# sheet.append()  整行写入数据到表格中, 括号内部需要传数据容器(元组, 列表)
sheet.append((1, 2, 3, 4, 5))
sheet.append((6, 7, 8, 9, 10))

# 保存此工作蒲对象
work_book.save('实例.xlsx')

"""爬虫课程只会用表格存数据"""
