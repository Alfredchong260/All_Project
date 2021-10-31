import openpyxl

'''创建工作簿对象'''
work_book = openpyxl.Workbook()
work = openpyxl.load_workbook('实例.xlsx')

'''创建对象'''
sheet = work_book.create_sheet('表1')

'''数据写入'''
# sheet['A1'] =  'A1'
# sheet['B1'] =  'B1'

# cell是单元格对象
# sheet.cell(row=1, column=1).value = '123456'
# sheet.cell(row=2, column=1).value = '222222'

# 整行写入数据到表格中  接受元祖或者列表
# sheet.append((1,2,3,4,5,6,7,8,9,0))
# sheet.append([ 0,1,2,3,4,5,6,7,8,9 ])

# 读取文件
for i in range(1, 10):
    for j in range(1, i + 1):
        # print(f'{j} * {i} = {i*j}', end='\t')
        sheet.cell(row=i, column=j).value = f'{j} * {i} = {i*j}'
    print()

# 获取表的名字
print(work.sheetnames)

sheet1 = work['表1']
# 读取最大行
# print(sheet1.max_row)
# 读取最大列
# print(sheet1.max_column)

# for i in range(1, sheet1.max_column + 1):
#     print(sheet1.cell(row=1, column=i).value)

# for i in range(1, sheet1.max_row + 1):
#     print(sheet1.cell(row=i, column=1).value)

for i in range(1, sheet1.max_row + 1):
    for j in range(1, sheet1.max_column + 1):
        print(sheet1.cell(row=i, column=j).value)
    print()

work_book.save('实例.xlsx')
